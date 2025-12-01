from openpyxl import load_workbook

from internet.models import *
from traceback import print_exc


DISCOUNT_TYPE_DICT = {
    "M": "Mobile",
    "I": "Internet",
    "TV": "TV",
    "I_I": "Internet Install",
    "TV_I": "TV Install",
    "W": "Wifi",
}


def parse_integer(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def get_or_create_bundle_promotion(bc_id, coupon_amount, cash_amount):
    if coupon_amount == 0 and cash_amount == 0:
        return

    try:
        bp, created = BundlePromotion.objects.update_or_create(
            bundle_condition_id=bc_id,
            defaults={
                "coupon_amount": coupon_amount,
                "cash_amount": cash_amount,
            },
            create_defaults={
                "coupon_amount": coupon_amount,
                "cash_amount": cash_amount,
            },
        )
    except Exception:
        print(
            "Error in get_or_create_bundle_promotion:",
            bc_id,
            coupon_amount,
            cash_amount,
        )
        print_exc()
        breakpoint()


def get_or_create_bundle_discount(bc_id, text):
    if not text:
        return

    discounts = text.strip(",").replace(",,", ",").replace("),", ")|").split("|")
    for disc in discounts:
        try:
            if disc[0] != "(" or disc[-1] != ")":
                raise ValueError("Invalid discount format:", disc)
            name, disc_type, price = disc.strip()[1:-1].split(",")
            price = parse_integer(price)
            if DISCOUNT_TYPE_DICT.get(disc_type.strip()) is None:
                raise ValueError("Invalid discount type:", disc_type)

            bd, created = BundleDiscount.objects.get_or_create(
                bundle_condition_id=bc_id,
                bundle_name=name.strip(),
                discount_type=DISCOUNT_TYPE_DICT.get(disc_type.strip()),
                discount_amount=price,
            )
        except Exception:
            print("Error in get_or_create_bundle_discount:", disc)
            print_exc()
            breakpoint()


def import_internet_plan_from_excel():
    wb = load_workbook("internet_tv_all.xlsx")
    ws = wb.active

    carrier_col = 1  # A
    internet_id_col = 6  # F
    wifi_id_col = 7  # G
    tv_id_col = 8  # H
    settop_id_col = 9  # I
    mobile_type_col = 10  # J
    etc_discount_col = 14  # N
    install_discount_col = 15  # O
    month_discount_col = 16  # P
    coupon_col = 17  # Q
    cash_col = 18  # R

    carriers = InternetCarrier.objects.all()
    internet_plans = InternetPlan.objects.select_related("carrier").all()
    wifi_options = WifiOption.objects.all()
    tv_plans = TVPlan.objects.all()
    settop_box_options = SettopBoxOption.objects.all()
    carrier_dict = {c.name: c for c in carriers}
    internet_plan_dict = {ip.id: ip for ip in internet_plans}
    wifi_option_dict = {wo.id: wo for wo in wifi_options}
    tv_plan_dict = {tp.id: tp for tp in tv_plans}
    settop_box_option_dict = {so.id: so for so in settop_box_options}

    print("#######Starting import######")

    for row in ws.iter_rows(min_row=2, max_row=129):  # Skip header row
        carrier_name = row[carrier_col - 1].value
        internet_plan_id = parse_integer(row[internet_id_col - 1].value)
        wifi_option_id = parse_integer(row[wifi_id_col - 1].value) or None
        tv_plan_id = parse_integer(row[tv_id_col - 1].value) or None
        settop_box_option_id = parse_integer(row[settop_id_col - 1].value) or None
        mobile_type = row[mobile_type_col - 1].value
        mobile_type = mobile_type if mobile_type in ["MNO", "MVNO"] else None

        # (name, type, price),(name, type, price),(name, type, price) 가 텍스트로 하드코딩된 데이터
        etc_discount = row[etc_discount_col - 1].value or None
        install_discount = row[install_discount_col - 1].value or None
        month_discount = row[month_discount_col - 1].value or None

        carrier = carrier_dict.get(carrier_name)
        internet_plan = internet_plan_dict.get(internet_plan_id)
        wifi_option = (
            wifi_option_dict.get(wifi_option_id) if wifi_option_id is not None else None
        )
        tv_plan = tv_plan_dict.get(tv_plan_id)
        settop_box_option = (
            settop_box_option_dict.get(settop_box_option_id)
            if settop_box_option_id is not None
            else None
        )

        try:
            bc = BundleCondition.objects.get(
                carrier=carrier,
                internet_plan=internet_plan,
                tv_plan=tv_plan,
                wifi_option=wifi_option,
                settop_box_option=settop_box_option,
                mobile_type=mobile_type,
            )

        except BundleCondition.DoesNotExist:
            obj = BundleCondition(
                carrier=carrier,
                internet_plan=internet_plan,
                tv_plan=tv_plan,
                wifi_option=wifi_option,
                settop_box_option=settop_box_option,
                mobile_type=mobile_type,
                mobile_price_min=0,
            )
            obj.save()

        bc_id = bc.id if "bc" in locals() else obj.id
        print(f"NOW ::: {row[0]}, {bc_id}")

        get_or_create_bundle_discount(bc_id, etc_discount)
        get_or_create_bundle_discount(bc_id, install_discount)
        get_or_create_bundle_discount(bc_id, month_discount)

        get_or_create_bundle_promotion(
            bc_id,
            parse_integer(row[coupon_col - 1].value),
            parse_integer(row[cash_col - 1].value),
        )

    return
