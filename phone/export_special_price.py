from io import BytesIO
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone

from phone.models import Product, ProductOption, Inventory
from phone.constants import CarrierChoices, DiscountTypeChoices, ContractTypeChoices

CARRIER_ORDER = [CarrierChoices.SK, CarrierChoices.KT, CarrierChoices.LG]
CARRIER_DISPLAY = {
    CarrierChoices.SK: "SKT",
    CarrierChoices.KT: "KT",
    CarrierChoices.LG: "LGU+",
}
CONTRACT_TYPES = [ContractTypeChoices.MNP, ContractTypeChoices.CHANGE]
CONTRACT_SHORT = {ContractTypeChoices.MNP: "번호이동", ContractTypeChoices.CHANGE: "기기변경"}
PRODUCTS_PER_ROW = 4  # 특가 그리드에서 한 줄당 상품 수
MERGE_SPAN = 2  # 병합 열 수

# ── Styles ──
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(bold=True, size=14)
TITLE_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

BANNER_FONT = Font(bold=True, size=12, color="FFFFFF")
BANNER_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

BRAND_FONT = Font(bold=True, size=11)
BRAND_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

CARRIER_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
CARRIER_FILLS = {
    CarrierChoices.SK: PatternFill(
        start_color="E74C3C", end_color="E74C3C", fill_type="solid"
    ),
    CarrierChoices.KT: PatternFill(
        start_color="2E86C1", end_color="2E86C1", fill_type="solid"
    ),
    CarrierChoices.LG: PatternFill(
        start_color="E91E8C", end_color="E91E8C", fill_type="solid"
    ),
}


def _build_data():
    """활성 상품 + 재고 + 옵션 데이터를 수집하여 엑셀에 필요한 구조를 반환한다."""

    # 1. 활성 상품
    products = list(
        Product.objects.filter(
            is_active=True,
            best_price_option_id__isnull=False,
            deleted_at__isnull=True,
        )
        .select_related("device")
        .order_by("-sort_order")
    )
    if not products:
        return None

    product_map = {p.id: p for p in products}
    product_ids = [p.id for p in products]
    device_ids = list({p.device_id for p in products})

    # 2. 재고 → (device_id, carrier, storage_capacity) set
    in_stock_set = set()
    for inv in Inventory.objects.filter(
        device_variant__device_id__in=device_ids,
        count__gt=0,
        deleted_at__isnull=True,
    ).select_related("dealership", "device_variant"):
        in_stock_set.add(
            (
                inv.device_variant.device_id,
                inv.dealership.carrier,
                inv.device_variant.storage_capacity,
            )
        )

    # 3. 공시지원금 옵션 전체
    all_options = list(
        ProductOption.objects.filter(
            product_id__in=product_ids,
            discount_type=DiscountTypeChoices.SUBSIDY,
            deleted_at__isnull=True,
            plan__deleted_at__isnull=True,
        ).select_related("plan", "device_variant")
    )

    # 4. 재고 있는 옵션만
    product_device = {p.id: p.device_id for p in products}
    stock_options = [
        opt
        for opt in all_options
        if (
            product_device[opt.product_id],
            opt.plan.carrier,
            opt.device_variant.storage_capacity,
        )
        in in_stock_set
    ]

    # 5. (product_id, carrier)별 가장 비싼 요금제 가격
    max_plan_price: dict[tuple, int] = {}
    for opt in stock_options:
        key = (opt.product_id, opt.plan.carrier)
        if key not in max_plan_price or opt.plan.price > max_plan_price[key]:
            max_plan_price[key] = opt.plan.price

    # 6. 해당 요금제의 옵션만 선택
    option_map: dict[tuple, ProductOption] = {}
    for opt in stock_options:
        if opt.plan.price != max_plan_price.get((opt.product_id, opt.plan.carrier)):
            continue
        key = (
            opt.product_id,
            opt.plan.carrier,
            opt.device_variant.storage_capacity,
            opt.contract_type,
        )
        if key not in option_map:
            option_map[key] = opt

    # 7. 상품 목록: (product_id, storage_capacity)
    items_set = set()
    for pid, _carrier, storage, _ctype in option_map:
        items_set.add((pid, storage))

    items = sorted(
        items_set,
        key=lambda ps: (-product_map[ps[0]].sort_order, ps[1]),
    )

    if not items:
        return None

    return product_map, option_map, items, max_plan_price


def _cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment or CENTER
    if border is not False:
        cell.border = border or THIN_BORDER
    return cell


def _merge_cell(ws, row, col_start, col_span, value, font=None, fill=None):
    """col_span개 열을 병합하고 값+스타일을 적용한다."""
    col_end = col_start + col_span - 1
    ws.merge_cells(
        start_row=row, start_column=col_start, end_row=row, end_column=col_end
    )
    _cell(ws, row, col_start, value, font, fill)
    # 병합된 나머지 셀에도 테두리/배경 적용
    for c in range(col_start + 1, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def _item_label(product_map, items, pid, storage):
    """상품 라벨. 같은 상품의 용량이 여러개면 용량도 표시."""
    product = product_map[pid]
    storages_for_product = [s for (p, s) in items if p == pid]
    if len(storages_for_product) > 1:
        return f"{product.name} {storage}"
    return product.name


def _price_만(final_price):
    """final_price(원)를 만원 단위 정수로 변환."""
    if final_price is None:
        return None
    return round(final_price / 10000)


def _write_banner(ws, row, col_end):
    """인터넷+TV 배너 행 작성."""
    _merge_cell(
        ws, row, 1, col_end,
        "인터넷+TV !!! 가입시 최대지원 !!",
        BANNER_FONT, BANNER_FILL,
    )
    ws.row_dimensions[row].height = 28
    return row + 1


SPECIAL_PRICE_LIMIT = 200000  # 특가폰 기준: final_price 20만원 이하


def _get_best_mnp(option_map, pid, storage):
    """번호이동 기준 최저가 옵션과 해당 통신사를 반환한다."""
    best_price = None
    best_carrier = None
    for carrier in CARRIER_ORDER:
        key = (pid, carrier, storage, ContractTypeChoices.MNP)
        opt = option_map.get(key)
        if opt and (best_price is None or opt.final_price < best_price):
            best_price = opt.final_price
            best_carrier = carrier
    return best_price, best_carrier


def generate_special_price_excel():
    """활성 상품의 특가폰 모음 엑셀을 생성하여 BytesIO로 반환한다."""

    data = _build_data()
    if data is None:
        return None

    product_map, option_map, items, max_plan_price = data

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "특가폰 모음"

    now = timezone.localtime(timezone.now())
    month = now.month
    today_str = now.strftime("%Y-%m-%d")

    # 상단 그리드: 상품당 2열 × 4개 = 8열
    grid_total_cols = PRODUCTS_PER_ROW * MERGE_SPAN
    # 하단 테이블: 제품명(2열) + 3통신사 × 2(번이/기변) = 8열
    table_total_cols = MERGE_SPAN + len(CARRIER_ORDER) * len(CONTRACT_TYPES)

    row = 1

    # ═══════════════════════════════════════════
    # Section 1: 특가폰 모음 (포스터 스타일 그리드)
    #   각 상품 = 2열 병합
    # ═══════════════════════════════════════════

    # 타이틀
    _merge_cell(
        ws, row, 1, grid_total_cols,
        f"{month}월 특가폰 모음", TITLE_FONT, TITLE_FILL,
    )
    ws.row_dimensions[row].height = 30
    row += 1

    # 특가 대상: 번호이동 최저가 20만원 이하인 상품만
    special_items = []
    for pid, storage in items:
        best_price, _carrier = _get_best_mnp(option_map, pid, storage)
        if best_price is not None and best_price <= SPECIAL_PRICE_LIMIT:
            special_items.append((pid, storage))

    for chunk_start in range(0, len(special_items), PRODUCTS_PER_ROW):
        chunk = special_items[chunk_start : chunk_start + PRODUCTS_PER_ROW]

        # 상품명 행 (2열 병합)
        for ci, (pid, storage) in enumerate(chunk):
            label = _item_label(product_map, items, pid, storage)
            col = 1 + ci * MERGE_SPAN
            _merge_cell(ws, row, col, MERGE_SPAN, label, Font(bold=True, size=9), HEADER_FILL)
        row += 1

        # 통신사 행 (어느 통신사 번호이동인지, 2열 병합)
        for ci, (pid, storage) in enumerate(chunk):
            col = 1 + ci * MERGE_SPAN
            _best_price, best_carrier = _get_best_mnp(option_map, pid, storage)
            if best_carrier:
                _merge_cell(
                    ws, row, col, MERGE_SPAN,
                    f"{CARRIER_DISPLAY[best_carrier]} 번호이동",
                    Font(bold=True, size=9, color="FFFFFF"),
                    CARRIER_FILLS[best_carrier],
                )
            else:
                _merge_cell(ws, row, col, MERGE_SPAN, "-", Font(size=9, color="999999"))
        row += 1

        # 특가 가격 행 (큰 폰트, 2열 병합)
        for ci, (pid, storage) in enumerate(chunk):
            col = 1 + ci * MERGE_SPAN
            best_price, _carrier = _get_best_mnp(option_map, pid, storage)

            price_fill = PatternFill(
                start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
            )
            if best_price is not None:
                val = _price_만(best_price)
                font = Font(
                    bold=True, size=22,
                    color="0000FF" if val < 0 else "FF0000",
                )
                _merge_cell(ws, row, col, MERGE_SPAN, val, font, price_fill)
            else:
                _merge_cell(ws, row, col, MERGE_SPAN, "-", Font(size=22), price_fill)

        ws.row_dimensions[row].height = 40
        row += 1

    row += 1  # 간격

    # ═══════════════════════════════════════════
    # Section 2: 브랜드별 상세 테이블
    #   제품명(2열 병합) | SKT 번이 | SKT 기변 | KT 번이 | KT 기변 | LG 번이 | LG 기변
    # ═══════════════════════════════════════════

    row = _write_banner(ws, row, table_total_cols)
    row += 1

    # 날짜 (2열 병합)
    _merge_cell(ws, row, 1, MERGE_SPAN, today_str, HEADER_FONT)
    row += 1

    # 브랜드별 그룹핑
    brand_groups = defaultdict(list)
    for pid, storage in items:
        brand = product_map[pid].device.brand
        brand_groups[brand].append((pid, storage))

    for brand, brand_items in brand_groups.items():

        # 브랜드 헤더
        _merge_cell(
            ws, row, 1, table_total_cols,
            f"{brand} 제품군", BRAND_FONT, BRAND_FILL,
        )
        row += 1

        # 통신사 헤더 (각 통신사 2열 병합) + 요금제 가격 표시
        _merge_cell(
            ws, row, 1, MERGE_SPAN,
            "공시지원금 개통기준\n요금제 188일 유지",
            HEADER_FONT, HEADER_FILL,
        )
        for ci, carrier in enumerate(CARRIER_ORDER):
            col = MERGE_SPAN + 1 + ci * len(CONTRACT_TYPES)
            # 해당 브랜드 상품들의 요금제 가격 수집
            plan_prices = set()
            for pid, _storage in brand_items:
                pp = max_plan_price.get((pid, carrier))
                if pp:
                    plan_prices.add(pp)
            price_text = "/".join(f"{p:,}원" for p in sorted(plan_prices))
            carrier_label = (
                f"{CARRIER_DISPLAY[carrier]}\n({price_text})" if price_text
                else CARRIER_DISPLAY[carrier]
            )
            _merge_cell(
                ws, row, col, len(CONTRACT_TYPES),
                carrier_label,
                CARRIER_HEADER_FONT, CARRIER_FILLS[carrier],
            )
        ws.row_dimensions[row].height = 36
        row += 1

        # 약정유형 서브헤더 (제품명 2열 병합)
        _merge_cell(ws, row, 1, MERGE_SPAN, "제품명", HEADER_FONT, HEADER_FILL)
        for ci, carrier in enumerate(CARRIER_ORDER):
            for cti, contract_type in enumerate(CONTRACT_TYPES):
                col = MERGE_SPAN + 1 + ci * len(CONTRACT_TYPES) + cti
                _cell(
                    ws, row, col,
                    CONTRACT_SHORT[contract_type],
                    Font(bold=True, size=9), HEADER_FILL,
                )
        row += 1

        # 상품 데이터 행 (제품명 2열 병합)
        for pid, storage in brand_items:
            label = _item_label(product_map, items, pid, storage)
            _merge_cell(ws, row, 1, MERGE_SPAN, label, Font(size=10))

            for ci, carrier in enumerate(CARRIER_ORDER):
                for cti, contract_type in enumerate(CONTRACT_TYPES):
                    col = MERGE_SPAN + 1 + ci * len(CONTRACT_TYPES) + cti
                    key = (pid, carrier, storage, contract_type)
                    opt = option_map.get(key)

                    cell = ws.cell(row=row, column=col)
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                    if opt:
                        val = _price_만(opt.final_price)
                        cell.value = val
                        cell.font = (
                            Font(bold=True, color="0000FF")
                            if val < 0
                            else Font()
                        )
                    else:
                        cell.value = "재고소진"
                        cell.font = Font(size=8, color="999999")
            row += 1

        row += 1  # 브랜드 간 간격

    # ═══════════════════════════════════════════
    # 컬럼 너비 조정
    # ═══════════════════════════════════════════
    for ci in range(max(grid_total_cols, table_total_cols)):
        ws.column_dimensions[get_column_letter(ci + 1)].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
