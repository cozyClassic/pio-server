import openpyxl
import io
from ..models import ProductOption

HEADERS = {
    "C": "5G 프리미어 슈퍼_번호이동_공시지원금",
    "D": "5G 프리미어 슈퍼_기기변경_공시지원금",
    "E": "5G 프리미어 슈퍼_번호이동_선택약정",
    "F": "5G 프리미어 슈퍼_기기변경_선택약정",
    "G": "5G 프리미어 플러스_번호이동_공시지원금",
    "H": "5G 프리미어 플러스_기기변경_공시지원금",
    "I": "5G 프리미어 플러스_번호이동_선택약정",
    "J": "5G 프리미어 플러스_기기변경_선택약정",
    "K": "5G 프리미어 레귤러_번호이동_공시지원금",
    "L": "5G 프리미어 레귤러_기기변경_공시지원금",
    "M": "5G 프리미어 레귤러_번호이동_선택약정",
    "N": "5G 프리미어 레귤러_기기변경_선택약정",
}

PLAN_START_COL = HEADERS.keys().__iter__().__next__()
PLAN_START_ROW = 3
PLAN_END_ROW = 18
PRICE_UNIT = 10000
MODEL_NAME_COL = "A"


def update_product_option_LG_subsidy_addtional(file: bytes, margin=0) -> None:
    """기능
    - 공시지원금은 별도로 관리하기 위해 여기에서 관리하지 않는다.
    1. product option을 db로부터 읽어온다.
    - 요금제의 통신사가 LG인것만, 삭제되지 않은것만
    2. KEY = 요금제명_약정유형_가입유형_lg명
      -> 요금제명 // plan
      -> 약정유형,가입유형 // product option
      -> lg명 // device variant
    3. 엑셀의 각 행을 읽어온다.
      -> A열 -> model_lg, C~N열 -> 요금제명_약정유형_가입유형
    4. 이미 db에 존재하는 경우 추가지원금을 업데이트 한다.
    5. db에 존재하지 않는 경우에는 별도로 추가하지 않는다. (관리할 요금제를 최소화하기 위함)
    """

    wb = openpyxl.load_workbook(io.BytesIO(file), data_only=True)
    ws = wb.active

    db_options = (
        ProductOption.objects.select_related(
            "plan", "device_variant", "device_variant__device"
        )
        .filter(
            deleted_at__isnull=True,
            plan__carrier="LG",
        )
        .exclude(device_variant__name_lg="")
    )

    db_option_dict = {}
    updates = []
    update_device_variants = set()

    for option in db_options:
        # name_lg가 일치하는 경우가 있음 (용량 여러개에 정책파일 row 1개)
        key = f"{option.plan.name}_{option.contract_type}_{option.discount_type}_{option.device_variant.name_lg}"
        if key in db_option_dict:
            db_option_dict[key].append(option)
        else:
            db_option_dict[key] = [option]

    for i in range(PLAN_START_ROW, PLAN_END_ROW + 1):
        model_name = ws[f"{MODEL_NAME_COL}{i}"].value
        if not model_name:
            continue

        for col, header in HEADERS.items():
            jungchaek = max(int(ws[f"{col}{i}"].value) * PRICE_UNIT - margin, 0)
            key = f"{header}_{model_name}"
            if key in db_option_dict:
                options = db_option_dict[key]
                for option in options:
                    option.additional_discount = jungchaek
                    option.final_price = option._get_final_price()
                    if option.final_price < 0:
                        option.additional_discount += option.final_price
                        option.final_price = 0
                    updates.append(option)
                    update_device_variants.add(
                        option.device_variant.device.model_name
                        + "_"
                        + option.device_variant.storage_capacity
                    )

    ProductOption.objects.bulk_update(updates, ["additional_discount", "final_price"])
    update_device_variants = sorted(list(update_device_variants))

    return f"{ws.title} 시트의 {update_device_variants}의 LG 추가지원금 {len(updates)}건 업데이트 완료"
