import openpyxl
import io
from ..models import ProductOption


HEADERS = {
    "G": "5GX 프리미엄_번호이동_공시지원금",
    "H": "5GX 프리미엄_번호이동_선택약정",
    "I": "5GX 프리미엄_기기변경_공시지원금",
    "J": "5GX 프리미엄_기기변경_선택약정",
    "L": "5GX 프라임플러스_번호이동_공시지원금",
    "M": "5GX 프라임플러스_번호이동_선택약정",
    "N": "5GX 프라임플러스_기기변경_공시지원금",
    "O": "5GX 프라임플러스_기기변경_선택약정",
    "Q": "5GX 프라임_번호이동_공시지원금",
    "R": "5GX 프라임_번호이동_선택약정",
    "S": "5GX 프라임_기기변경_공시지원금",
    "T": "5GX 프라임_기기변경_선택약정",
    "V": "5GX 레귤러플러스_번호이동_공시지원금",
    "W": "5GX 레귤러플러스_번호이동_선택약정",
    "X": "5GX 레귤러플러스_기기변경_공시지원금",
    "Y": "5GX 레귤러플러스_기기변경_선택약정",
}

PLAN_START_COL = HEADERS.keys().__iter__().__next__()
PLAN_START_ROW = 9
PLAN_END_ROW = 57
PRICE_UNIT = 1000
MODEL_NAME_COL = "B"


def update_product_option_SK_subsidy_addtional(file: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(file))
    ws = wb.active

    """기능
    - 공시지원금은 별도로 관리하기 위해 여기에서 관리하지 않는다.
    1. product option을 db로부터 읽어온다.
    - 요금제의 통신사가 SK인것만, 삭제되지 않은것만
    2. KEY = 요금제명_약정유형_가입유형_sk명
      -> 요금제명 = plan
      -> 약정유형,가입유형 = product option
      -> sk명 = device variant
    3. 엑셀의 각 행을 읽어온다.
      -> B열 -> model_sk, G~Y열 -> 요금제명_약정유형_가입유형
    4. 이미 db에 존재하는 경우 추가지원금을 업데이트 한다.
    5. db에 존재하지 않는 경우에는 별도로 추가하지 않는다. (관리할 요금제를 최소화하기 위함)
    """

    db_options = (
        ProductOption.objects.select_related(
            "plan", "device_variant", "device_variant__device"
        )
        .filter(
            deleted_at__isnull=True,
            plan__carrier="SK",
        )
        .exclude(device_variant__name_sk="")
    )

    db_option_dict = {}
    updates = []
    update_device_variants = set()

    for option in db_options:
        # name_sk가 일치하는 경우가 있음 (ex. 아이폰16같은 경우에는 용량별로 정책이 동일해서 name_sk가 같음)
        key = f"{option.plan.name}_{option.contract_type}_{option.discount_type}_{option.device_variant.name_sk}"
        if key in db_option_dict:
            db_option_dict[key].append(option)
        else:
            db_option_dict[key] = [option]

    for i in range(PLAN_START_ROW, PLAN_END_ROW + 1):
        model_name = ws[f"{MODEL_NAME_COL}{i}"].value
        for col, header in HEADERS.items():
            jungchaek = int(ws[f"{col}{i}"].value) * PRICE_UNIT
            key = f"{header}_{model_name}"
            if key in db_option_dict:
                options = db_option_dict[key]
                for option in options:
                    option.additional_discount = jungchaek
                    updates.append(option)
                    update_device_variants.add(
                        option.device_variant.device.model_name
                        + "_"
                        + option.device_variant.storage_capacity
                    )

    ProductOption.objects.bulk_update(updates, ["additional_discount"])
    update_device_variants = sorted(list(update_device_variants))

    return f"{ws.title} 시트의 {update_device_variants}의 SK 추가지원금 {len(updates)}건 업데이트 완료"
