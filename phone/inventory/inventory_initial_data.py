import openpyxl


from phone.models import Inventory, Dealership, DeviceVariant, DeviceColor

"""
각 대리점에서 받는 재고표와 DB에 있는 device_varaint, device_color 정보를 매칭하기 위한 DB 초기 데이터 로드
"""

HEADERS = {
    "dv_id": "device_variant.id",
    "model_name": "device.model_name",
    "capacity": "device_variant.storage_capacity",
    "sk_name_sheet": "디아이_모델명",
    "sk_color_sheet": "디아이_색상",
    "kt_name_sheet": "퍼스트_모델명",
    "kt_color_sheet": "퍼스트_색상",
    "lg_name_sheet": "휴넷_모델명",
    "lg_color_sheet": "휴넷_색상",
    "color": "device_color.color",
    "color_id": "device_color.id",
}


def load_initial_inventory_data(file_path: str):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    dv_dict = {int(dv.id): dv for dv in DeviceVariant.objects.all()}
    color_dict = {int(dc.id): dc for dc in DeviceColor.objects.all()}
    dealers = {
        "sk": Dealership.objects.get(name="디아이"),
        "kt": Dealership.objects.get(name="퍼스트"),
        "lg": Dealership.objects.get(name="엘비휴넷"),
    }

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(HEADERS.keys(), row))

        # 매핑 설정 (대리점 코드 단축)
        mapping = [
            ("sk", "sk_color_sheet", "sk_name_sheet"),
            ("lg", "lg_color_sheet", "lg_name_sheet"),
            ("kt", "kt_color_sheet", "kt_name_sheet"),
        ]

    inventory_objects = []

    # 2. 엑셀을 돌면서 객체 리스트 생성 (DB 접근 X)
    try:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(HEADERS.keys(), row))

            if (
                row_data["dv_id"] is None
                or row_data["color_id"] is None
                or dv_dict.get(int(row_data["dv_id"])) is None
                or color_dict.get(int(row_data["color_id"])) is None
            ):
                continue

            mapping = [
                ("sk", "sk_color_sheet", "sk_name_sheet"),
                ("lg", "lg_color_sheet", "lg_name_sheet"),
                ("kt", "kt_color_sheet", "kt_name_sheet"),
            ]

            for d_key, color_key, name_key in mapping:
                if row_data[color_key] and row_data[name_key]:
                    # 객체 생성 (저장은 하지 않음)
                    inventory_objects.append(
                        Inventory(
                            device_variant=dv_dict[int(row_data["dv_id"])],
                            device_color=color_dict[int(row_data["color_id"])],
                            dealership=dealers[d_key],
                            count=0,
                            color_in_sheet=row_data[color_key].replace(" ", ""),
                            name_in_sheet=row_data[name_key],
                        )
                    )

        # 3. 한 번에 DB로 쏘기 (Bulk Upsert)
        if inventory_objects:
            Inventory.objects.bulk_create(
                inventory_objects,
                update_conflicts=True,
                unique_fields=[
                    "device_variant",
                    "device_color",
                    "dealership",
                ],  # 중복 기준 필드
                update_fields=[
                    "color_in_sheet",
                    "name_in_sheet",
                ],  # 이미 있을 때 업데이트할 필드
            )

    except Exception as e:
        print(f"Error occurred: {e}")
        breakpoint()
